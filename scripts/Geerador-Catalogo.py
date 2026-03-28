import csv
import os
import shutil
import subprocess
import sys
import unicodedata
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook


STATUS_ATIVO = {"ativo"}


def valor_status_ativo(valor) -> bool:
    if valor is None:
        return False
    return str(valor).strip().lower() in STATUS_ATIVO


def slugify(texto: str) -> str:
    texto = texto.strip().lower()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    permitido = []
    for c in texto:
        if c.isalnum() or c in (" ", "-", "_"):
            permitido.append(c)
    texto = "".join(permitido).replace("_", "-")
    texto = "-".join(texto.split())
    while "--" in texto:
        texto = texto.replace("--", "-")
    return texto.strip("-")


def limpar_descricao_linha_unica(texto: str) -> str:
    return " ".join(str(texto).replace("\r", " ").replace("\n", " ").split())


def abrir_no_sistema(caminho: Path):
    if not caminho.exists():
        raise FileNotFoundError(f"Caminho não encontrado:\n{caminho}")

    if sys.platform.startswith("win"):
        os.startfile(str(caminho))
    elif sys.platform == "darwin":
        subprocess.run(["open", str(caminho)], check=False)
    else:
        subprocess.run(["xdg-open", str(caminho)], check=False)


class GeradorCatalogoApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Additive Hub • Gerador de Catálogo")
        self.root.geometry("1360x860")
        self.root.minsize(1180, 720)

        try:
            self.root.state("zoomed")
        except Exception:
            pass

        self.imagens_selecionadas = []
        self.produtos_csv_cache = []

        self.var_pasta_projeto = tk.StringVar()
        self.var_status = tk.StringVar(value="ativo")
        self.var_id = tk.StringVar()
        self.var_nome_produto = tk.StringVar()
        self.var_categoria = tk.StringVar(value="Chaveiro")
        self.var_subcategoria = tk.StringVar()
        self.var_preco = tk.StringVar()
        self.var_destaque = tk.StringVar()
        self.var_pasta_produto = tk.StringVar()
        self.var_produto_existente = tk.StringVar()

        self.var_nome_variacao_1 = tk.StringVar()
        self.var_variacoes_1 = tk.StringVar()
        self.var_nome_variacao_2 = tk.StringVar()
        self.var_variacoes_2 = tk.StringVar()
        self.var_peso = tk.StringVar()
        self.var_altura = tk.StringVar()
        self.var_largura = tk.StringVar()
        self.var_comprimento = tk.StringVar()

        self.var_planilha_massa = tk.StringVar()
        self.var_pasta_fotos = tk.StringVar()

        self.categorias = {
            "Chaveiro": {
                "pasta": "chaveiros",
                "subcategorias": [
                    "Slim Classic",
                    "Slim Smart",
                    "Slim Personalizado",
                    "Slim Relevo",
                    "3D Basic",
                    "3D Classic",
                    "3D Spin",
                    "3D Mini-Flexi",
                ],
            },
            "Decoração": {
                "pasta": "decoracoes",
                "subcategorias": [
                    "Miniaturas",
                    "Itens para cozinha",
                    "Itens para ambiente",
                ],
            },
            "Cozinha & Confeitaria": {
                "pasta": "cozinha-confeitaria",
                "subcategorias": [
                    "Cortadores",
                    "Marcadores",
                    "Utensílios personalizados",
                ],
            },
            "Utilidades": {
                "pasta": "utilidades",
                "subcategorias": [
                    "Organizadores",
                    "Suportes",
                    "Acessórios funcionais",
                ],
            },
            "Personalizados": {
                "pasta": "personalizados",
                "subcategorias": [
                    "Projetos sob medida",
                    "Brindes personalizados",
                    "Peças exclusivas",
                ],
            },
        }

        self._configurar_estilo()
        self._montar_interface_com_scroll()
        self._atualizar_subcategorias()
        self.gerar_id_automatico()
        self.atualizar_pasta_produto_automatica()

    def _configurar_estilo(self):
        style = ttk.Style()
        try:
            if "vista" in style.theme_names():
                style.theme_use("vista")
        except Exception:
            pass

        style.configure("Title.TLabel", font=("Segoe UI", 18, "bold"))
        style.configure("Subtitle.TLabel", font=("Segoe UI", 10))
        style.configure("Card.TLabelframe", padding=10)
        style.configure("Card.TLabelframe.Label", font=("Segoe UI", 10, "bold"))
        style.configure("Accent.TButton", font=("Segoe UI", 9, "bold"))
        style.configure("Small.TButton", font=("Segoe UI", 9))

    def _montar_interface_com_scroll(self):
        container = ttk.Frame(self.root)
        container.pack(fill="both", expand=True)

        self.canvas = tk.Canvas(container, highlightthickness=0)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=scrollbar.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        def _on_mousewheel(event):
            self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        self.canvas.bind_all("<MouseWheel>", _on_mousewheel)

        self._montar_interface()

    def _montar_interface(self):
        frame = ttk.Frame(self.scrollable_frame, padding=14)
        frame.pack(fill="both", expand=True)

        topo = ttk.Frame(frame)
        topo.pack(fill="x", pady=(0, 10))

        ttk.Label(
            topo,
            text="Gerador de catálogo Additive Hub",
            style="Title.TLabel",
        ).pack(anchor="w")
        ttk.Label(
            topo,
            text="Cadastre produtos, organize imagens e monte o CSV com muito menos trabalho.",
            style="Subtitle.TLabel",
        ).pack(anchor="w", pady=(2, 0))

        projeto = ttk.LabelFrame(frame, text="Projeto", style="Card.TLabelframe")
        projeto.pack(fill="x", pady=(0, 10))

        linha_projeto = ttk.Frame(projeto)
        linha_projeto.pack(fill="x")

        ttk.Entry(linha_projeto, textvariable=self.var_pasta_projeto).pack(
            side="left", fill="x", expand=True, padx=(0, 8)
        )
        ttk.Button(
            linha_projeto,
            text="Selecionar pasta do projeto",
            style="Accent.TButton",
            command=self.selecionar_pasta_projeto,
        ).pack(side="left", padx=(0, 8))
        ttk.Button(
            linha_projeto,
            text="Abrir produtos.csv",
            style="Small.TButton",
            command=self.abrir_csv,
        ).pack(side="left")

        massa = ttk.LabelFrame(frame, text="Importação em massa", style="Card.TLabelframe")
        massa.pack(fill="x", pady=(0, 10))

        linha_massa_1 = ttk.Frame(massa)
        linha_massa_1.pack(fill="x", pady=(0, 6))
        ttk.Label(linha_massa_1, text="Planilha", width=18).pack(side="left")
        ttk.Entry(linha_massa_1, textvariable=self.var_planilha_massa).pack(
            side="left", fill="x", expand=True, padx=(0, 8)
        )
        ttk.Button(
            linha_massa_1,
            text="Selecionar CSV / Excel",
            command=self.selecionar_planilha_massa,
        ).pack(side="left")

        linha_massa_2 = ttk.Frame(massa)
        linha_massa_2.pack(fill="x", pady=(0, 6))
        ttk.Label(linha_massa_2, text="Pasta base das fotos", width=18).pack(side="left")
        ttk.Entry(linha_massa_2, textvariable=self.var_pasta_fotos).pack(
            side="left", fill="x", expand=True, padx=(0, 8)
        )
        ttk.Button(
            linha_massa_2,
            text="Selecionar pasta",
            command=self.selecionar_pasta_fotos,
        ).pack(side="left")

        linha_massa_3 = ttk.Frame(massa)
        linha_massa_3.pack(fill="x")
        ttk.Button(
            linha_massa_3,
            text="🚀 Processar importação em massa",
            style="Accent.TButton",
            command=self.processar_em_massa,
        ).pack(side="left")

        acoes = ttk.LabelFrame(frame, text="Ações rápidas", style="Card.TLabelframe")
        acoes.pack(fill="x", pady=(0, 10))

        barra_acoes_1 = ttk.Frame(acoes)
        barra_acoes_1.pack(fill="x", pady=(0, 6))
        barra_acoes_2 = ttk.Frame(acoes)
        barra_acoes_2.pack(fill="x")

        ttk.Button(
            barra_acoes_1,
            text="Gerar ID automático",
            command=self.gerar_id_automatico_e_pasta,
        ).pack(side="left", padx=(0, 8))
        ttk.Button(
            barra_acoes_1,
            text="Selecionar imagens",
            command=self.selecionar_imagens,
        ).pack(side="left", padx=(0, 8))
        ttk.Button(
            barra_acoes_1,
            text="Remover imagem",
            command=self.remover_imagem_selecionada,
        ).pack(side="left", padx=(0, 8))
        ttk.Button(
            barra_acoes_1,
            text="Processar imagens",
            command=self.processar_imagens,
        ).pack(side="left", padx=(0, 8))

        ttk.Button(
            barra_acoes_2,
            text="Gerar linha CSV",
            command=self.gerar_linha_csv,
        ).pack(side="left", padx=(0, 8))
        ttk.Button(
            barra_acoes_2,
            text="Adicionar ao produtos.csv",
            command=self.adicionar_ao_csv,
        ).pack(side="left", padx=(0, 8))
        ttk.Button(
            barra_acoes_2,
            text="Abrir pasta do produto",
            command=self.abrir_pasta_produto,
        ).pack(side="left", padx=(0, 8))
        ttk.Button(
            barra_acoes_2,
            text="Limpar",
            command=self.limpar_tudo,
        ).pack(side="right")

        self.label_status = ttk.Label(acoes, text="Pronto.")
        self.label_status.pack(anchor="w", pady=(8, 0))

        corpo = ttk.Frame(frame)
        corpo.pack(fill="both", expand=True)

        esquerda = ttk.LabelFrame(corpo, text="Cadastro do produto", style="Card.TLabelframe")
        esquerda.pack(side="left", fill="both", expand=True, padx=(0, 6))

        direita = ttk.LabelFrame(corpo, text="Resultado", style="Card.TLabelframe")
        direita.pack(side="left", fill="both", expand=True, padx=(6, 0))

        self._montar_cadastro(esquerda)
        self._montar_resultado(direita)

    def _montar_cadastro(self, parent):
        form = ttk.Frame(parent)
        form.pack(fill="both", expand=True)

        linha1 = ttk.Frame(form)
        linha1.pack(fill="x", pady=4)
        ttk.Label(linha1, text="Produto já cadastrado", width=20).pack(side="left")
        self.combo_produtos = ttk.Combobox(
            linha1,
            textvariable=self.var_produto_existente,
            state="readonly",
        )
        self.combo_produtos.pack(side="left", fill="x", expand=True)
        self.combo_produtos.bind("<<ComboboxSelected>>", self.carregar_produto_existente)

        linha_status = ttk.Frame(form)
        linha_status.pack(fill="x", pady=4)
        ttk.Label(linha_status, text="Status", width=20).pack(side="left")
        self.combo_status = ttk.Combobox(
            linha_status,
            textvariable=self.var_status,
            values=["ativo", "inativo"],
            state="readonly",
            width=18,
        )
        self.combo_status.pack(side="left", fill="x")

        linha2 = ttk.Frame(form)
        linha2.pack(fill="x", pady=4)
        ttk.Label(linha2, text="ID", width=20).pack(side="left")
        entry_id = ttk.Entry(linha2, textvariable=self.var_id, width=12)
        entry_id.pack(side="left", padx=(0, 8))
        entry_id.bind("<KeyRelease>", self.atualizar_pasta_produto_automatica)

        ttk.Label(linha2, text="Preço", width=10).pack(side="left")
        ttk.Entry(linha2, textvariable=self.var_preco).pack(side="left", fill="x", expand=True)

        linha3 = ttk.Frame(form)
        linha3.pack(fill="x", pady=4)
        ttk.Label(linha3, text="Nome do produto", width=20).pack(side="left")
        entry_nome = ttk.Entry(linha3, textvariable=self.var_nome_produto)
        entry_nome.pack(side="left", fill="x", expand=True)
        entry_nome.bind("<KeyRelease>", self.atualizar_pasta_produto_automatica)

        linha4 = ttk.Frame(form)
        linha4.pack(fill="x", pady=4)
        ttk.Label(linha4, text="Categoria", width=20).pack(side="left")
        self.combo_categoria = ttk.Combobox(
            linha4,
            textvariable=self.var_categoria,
            values=list(self.categorias.keys()),
            state="readonly",
            width=24,
        )
        self.combo_categoria.pack(side="left", fill="x", expand=True, padx=(0, 8))
        self.combo_categoria.bind("<<ComboboxSelected>>", self.ao_mudar_categoria)

        ttk.Label(linha4, text="Subcategoria", width=14).pack(side="left")
        self.combo_subcategoria = ttk.Combobox(
            linha4,
            textvariable=self.var_subcategoria,
            state="readonly",
            width=28,
        )
        self.combo_subcategoria.pack(side="left", fill="x", expand=True)

        linha5 = ttk.Frame(form)
        linha5.pack(fill="x", pady=4)
        ttk.Label(linha5, text="Destaque", width=20).pack(side="left")
        ttk.Entry(linha5, textvariable=self.var_destaque).pack(side="left", fill="x", expand=True)

        linha6 = ttk.Frame(form)
        linha6.pack(fill="x", pady=4)
        ttk.Label(linha6, text="Pasta do produto", width=20).pack(side="left")
        ttk.Entry(linha6, textvariable=self.var_pasta_produto).pack(side="left", fill="x", expand=True)

        linha7 = ttk.Frame(form)
        linha7.pack(fill="x", pady=4)
        ttk.Label(linha7, text="Nome variação 1", width=20).pack(side="left")
        ttk.Entry(linha7, textvariable=self.var_nome_variacao_1).pack(
            side="left", fill="x", expand=True, padx=(0, 8)
        )
        ttk.Label(linha7, text="Variações 1", width=14).pack(side="left")
        ttk.Entry(linha7, textvariable=self.var_variacoes_1).pack(
            side="left", fill="x", expand=True
        )

        linha8 = ttk.Frame(form)
        linha8.pack(fill="x", pady=4)
        ttk.Label(linha8, text="Nome variação 2", width=20).pack(side="left")
        ttk.Entry(linha8, textvariable=self.var_nome_variacao_2).pack(
            side="left", fill="x", expand=True, padx=(0, 8)
        )
        ttk.Label(linha8, text="Variações 2", width=14).pack(side="left")
        ttk.Entry(linha8, textvariable=self.var_variacoes_2).pack(
            side="left", fill="x", expand=True
        )

        linha9 = ttk.Frame(form)
        linha9.pack(fill="x", pady=4)
        ttk.Label(linha9, text="Peso (kg)", width=20).pack(side="left")
        ttk.Entry(linha9, textvariable=self.var_peso).pack(
            side="left", fill="x", expand=True, padx=(0, 8)
        )
        ttk.Label(linha9, text="Altura (cm)", width=14).pack(side="left")
        ttk.Entry(linha9, textvariable=self.var_altura).pack(
            side="left", fill="x", expand=True
        )

        linha10 = ttk.Frame(form)
        linha10.pack(fill="x", pady=4)
        ttk.Label(linha10, text="Largura (cm)", width=20).pack(side="left")
        ttk.Entry(linha10, textvariable=self.var_largura).pack(
            side="left", fill="x", expand=True, padx=(0, 8)
        )
        ttk.Label(linha10, text="Comprimento (cm)", width=14).pack(side="left")
        ttk.Entry(linha10, textvariable=self.var_comprimento).pack(
            side="left", fill="x", expand=True
        )

        ttk.Label(form, text="Descrição").pack(anchor="w", pady=(8, 4))
        self.txt_descricao = tk.Text(form, height=5, wrap="word")
        self.txt_descricao.pack(fill="x")

        ttk.Label(form, text="Imagens selecionadas").pack(anchor="w", pady=(10, 4))
        self.lista_imagens = tk.Listbox(form, height=14)
        self.lista_imagens.pack(fill="both", expand=True)

    def _montar_resultado(self, parent):
        frame = ttk.Frame(parent)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text="Pasta criada / usada").pack(anchor="w")
        self.txt_pasta_saida = tk.Text(frame, height=2, wrap="word")
        self.txt_pasta_saida.pack(fill="x", pady=(4, 8))

        ttk.Label(frame, text='Campo "imagens"').pack(anchor="w")
        self.txt_campo_imagens = tk.Text(frame, height=5, wrap="word")
        self.txt_campo_imagens.pack(fill="x", pady=(4, 6))

        linha_botoes = ttk.Frame(frame)
        linha_botoes.pack(fill="x", pady=(0, 8))
        ttk.Button(
            linha_botoes,
            text='Copiar campo "imagens"',
            command=self.copiar_campo_imagens,
        ).pack(side="left", padx=(0, 8))
        ttk.Button(
            linha_botoes,
            text="Copiar linha CSV",
            command=self.copiar_linha_csv,
        ).pack(side="left")

        ttk.Label(frame, text="Linha CSV completa").pack(anchor="w")
        self.txt_linha_csv = tk.Text(frame, height=7, wrap="word")
        self.txt_linha_csv.pack(fill="x", pady=(4, 8))

        ttk.Label(frame, text="Arquivos gerados / relatório").pack(anchor="w")
        self.txt_arquivos_gerados = tk.Text(frame, height=16, wrap="word")
        self.txt_arquivos_gerados.pack(fill="both", expand=True, pady=(4, 0))

    def _set_status(self, texto: str):
        self.label_status.config(text=texto)

    def _set_texto(self, widget, texto: str):
        widget.delete("1.0", tk.END)
        widget.insert("1.0", texto)

    def _obter_subpasta_categoria(self, categoria=None) -> str:
        categoria = (categoria or self.var_categoria.get()).strip()
        if categoria not in self.categorias:
            return "outros"
        return self.categorias[categoria]["pasta"]

    def _atualizar_subcategorias(self):
        categoria = self.var_categoria.get()
        lista = self.categorias.get(categoria, {}).get("subcategorias", [])
        self.combo_subcategoria["values"] = lista
        if lista:
            self.var_subcategoria.set(lista[0])
        else:
            self.var_subcategoria.set("")

    def ao_mudar_categoria(self, _event=None):
        self._atualizar_subcategorias()

    def montar_nome_pasta_produto(self, id_produto: str, nome_produto: str) -> str:
        id_produto = str(id_produto).strip()
        nome_slug = slugify(nome_produto)

        if id_produto and nome_slug:
            return f"{id_produto}-{nome_slug}"
        if id_produto:
            return id_produto
        if nome_slug:
            return nome_slug
        return ""

    def atualizar_pasta_produto_automatica(self, _event=None):
        pasta = self.montar_nome_pasta_produto(
            self.var_id.get().strip(),
            self.var_nome_produto.get().strip()
        )
        self.var_pasta_produto.set(pasta)

    def obter_caminho_csv(self) -> Path:
        pasta_projeto = self.var_pasta_projeto.get().strip()
        if not pasta_projeto:
            raise ValueError("Selecione a pasta do projeto.")

        caminho_csv = Path(pasta_projeto) / "backend" / "data" / "produtos.csv"
        caminho_csv.parent.mkdir(parents=True, exist_ok=True)
        return caminho_csv

    def obter_base_imagens_backend(self) -> Path:
        pasta_projeto = self.var_pasta_projeto.get().strip()
        if not pasta_projeto:
            raise ValueError("Selecione a pasta do projeto.")

        base_imagens = Path(pasta_projeto) / "backend" / "data" / "imagens" / "produtos"
        base_imagens.mkdir(parents=True, exist_ok=True)
        return base_imagens

    def montar_caminho_relativo_imagem_backend(self, subpasta_categoria: str, pasta_produto: str, nome_arquivo: str) -> str:
        return f"backend/data/imagens/produtos/{subpasta_categoria}/{pasta_produto}/{nome_arquivo}"

    def selecionar_pasta_projeto(self):
        pasta = filedialog.askdirectory(title="Selecione a pasta do projeto")
        if not pasta:
            return
        self.var_pasta_projeto.set(pasta)
        self._set_status("Pasta do projeto selecionada.")
        self.carregar_produtos_csv()

    def selecionar_planilha_massa(self):
        arquivo = filedialog.askopenfilename(
            title="Selecione a planilha",
            filetypes=[
                ("Planilhas Excel e CSV", "*.xlsx *.csv"),
                ("Excel", "*.xlsx"),
                ("CSV", "*.csv"),
                ("Todos os arquivos", "*.*"),
            ],
        )
        if not arquivo:
            return

        self.var_planilha_massa.set(str(arquivo).strip())
        self._set_status("Planilha de importação em massa selecionada.")

    def selecionar_pasta_fotos(self):
        pasta = filedialog.askdirectory(title="Selecione a pasta base das fotos")
        if not pasta:
            return
        self.var_pasta_fotos.set(pasta)
        self._set_status("Pasta base das fotos selecionada.")

    def selecionar_imagens(self):
        arquivos = filedialog.askopenfilenames(
            title="Selecione as imagens do produto",
            filetypes=[
                ("Imagens", "*.png *.jpg *.jpeg *.webp"),
                ("Todos os arquivos", "*.*"),
            ],
        )
        if not arquivos:
            return

        self.imagens_selecionadas.extend(Path(a) for a in arquivos)

        vistos = set()
        unicos = []
        for p in self.imagens_selecionadas:
            chave = str(p.resolve())
            if chave not in vistos:
                vistos.add(chave)
                unicos.append(p)
        self.imagens_selecionadas = unicos

        self.atualizar_lista_imagens()
        self._set_status(f"{len(self.imagens_selecionadas)} imagem(ns) selecionada(s).")

    def atualizar_lista_imagens(self):
        self.lista_imagens.delete(0, tk.END)
        for i, arquivo in enumerate(self.imagens_selecionadas, start=1):
            self.lista_imagens.insert(tk.END, f"{i}. {arquivo.name}")

    def remover_imagem_selecionada(self):
        selecao = self.lista_imagens.curselection()
        if not selecao:
            messagebox.showwarning("Aviso", "Selecione uma imagem da lista para remover.")
            return
        indice = selecao[0]
        self.imagens_selecionadas.pop(indice)
        self.atualizar_lista_imagens()
        self._set_status("Imagem removida da seleção.")

    def _obter_pasta_destino(self) -> Path:
        pasta_produto = self.var_pasta_produto.get().strip()
        subpasta_categoria = self._obter_subpasta_categoria()

        if not self.var_pasta_projeto.get().strip():
            raise ValueError("Selecione a pasta do projeto.")
        if not pasta_produto:
            raise ValueError("Informe a pasta do produto.")

        return self.obter_base_imagens_backend() / subpasta_categoria / pasta_produto

    def processar_imagens(self):
        try:
            if not self.imagens_selecionadas:
                raise ValueError("Selecione pelo menos uma imagem.")

            destino = self._obter_pasta_destino()
            destino.mkdir(parents=True, exist_ok=True)

            subpasta_categoria = self._obter_subpasta_categoria()
            pasta_produto = self.var_pasta_produto.get().strip()

            caminhos_csv = []
            arquivos_gerados = []

            for i, origem in enumerate(self.imagens_selecionadas, start=1):
                extensao = origem.suffix.lower() if origem.suffix else ".png"
                novo_nome = f"{i}{extensao}"
                destino_arquivo = destino / novo_nome
                shutil.copy2(origem, destino_arquivo)

                caminho_relativo = self.montar_caminho_relativo_imagem_backend(
                    subpasta_categoria,
                    pasta_produto,
                    novo_nome,
                )
                caminhos_csv.append(caminho_relativo)
                arquivos_gerados.append(f"{origem.name}  ->  {novo_nome}")

            self._set_texto(self.txt_pasta_saida, str(destino))
            self._set_texto(self.txt_campo_imagens, "|".join(caminhos_csv))
            self._set_texto(self.txt_arquivos_gerados, "\n".join(arquivos_gerados))
            self._set_status("Imagens copiadas, organizadas e renomeadas com sucesso.")

        except Exception as e:
            messagebox.showerror("Erro", str(e))

    def gerar_id_automatico(self):
        try:
            pasta_projeto = self.var_pasta_projeto.get().strip()
            if not pasta_projeto:
                self.var_id.set("1")
                return

            caminho_csv = self.obter_caminho_csv()
            if not caminho_csv.exists():
                self.var_id.set("1")
                self._set_status("ID automático definido como 1.")
                return

            maior = 0
            with open(caminho_csv, "r", newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    valor_id = str(row.get("id", "")).strip()
                    if not valor_id:
                        continue

                    try:
                        valor = int(valor_id)
                    except ValueError:
                        numeros = "".join(ch for ch in valor_id if ch.isdigit())
                        if not numeros:
                            continue
                        valor = int(numeros)

                    if valor > maior:
                        maior = valor

            self.var_id.set(str(maior + 1))
            self._set_status(f"ID automático definido como {maior + 1}.")
        except Exception as e:
            messagebox.showerror("Erro", str(e))

    def gerar_id_automatico_e_pasta(self):
        self.gerar_id_automatico()
        self.atualizar_pasta_produto_automatica()

    def gerar_linha_csv(self):
        try:
            campo_imagens = self.txt_campo_imagens.get("1.0", tk.END).strip()
            if not campo_imagens:
                raise ValueError("Processe as imagens antes de gerar a linha CSV.")

            descricao = limpar_descricao_linha_unica(
                self.txt_descricao.get("1.0", tk.END)
            )

            linha = [
                self.var_status.get().strip() or "ativo",
                self.var_id.get().strip(),
                self.var_nome_produto.get().strip(),
                self.var_categoria.get().strip(),
                self.var_subcategoria.get().strip(),
                self.var_preco.get().strip(),
                self.var_destaque.get().strip(),
                descricao,
                campo_imagens,
                self.var_nome_variacao_1.get().strip(),
                self.var_variacoes_1.get().strip(),
                self.var_nome_variacao_2.get().strip(),
                self.var_variacoes_2.get().strip(),
                self.var_peso.get().strip(),
                self.var_altura.get().strip(),
                self.var_largura.get().strip(),
                self.var_comprimento.get().strip(),
            ]

            temp_path = Path("temp_linha.csv")

            with open(temp_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(
                    f,
                    delimiter=",",
                    quotechar='"',
                    quoting=csv.QUOTE_ALL,
                )
                writer.writerow(linha)

            with open(temp_path, "r", encoding="utf-8-sig") as f:
                buffer = f.read().strip()

            temp_path.unlink(missing_ok=True)

            self._set_texto(self.txt_linha_csv, buffer)
            self._set_status("Linha CSV gerada com sucesso.")
        except Exception as e:
            messagebox.showerror("Erro", str(e))

    def adicionar_ao_csv(self):
        try:
            pasta_projeto = self.var_pasta_projeto.get().strip()
            if not pasta_projeto:
                raise ValueError("Selecione a pasta do projeto.")

            campo_imagens = self.txt_campo_imagens.get("1.0", tk.END).strip()
            if not campo_imagens:
                raise ValueError("Processe as imagens antes de adicionar ao CSV.")

            caminho_csv = self.obter_caminho_csv()
            descricao = limpar_descricao_linha_unica(
                self.txt_descricao.get("1.0", tk.END)
            )

            cabecalho = [
                "status",
                "id",
                "nome",
                "categoria",
                "subcategoria",
                "preco",
                "destaque",
                "descricao",
                "imagens",
                "nome_variacao_1",
                "variacoes_1",
                "nome_variacao_2",
                "variacoes_2",
                "peso",
                "altura",
                "largura",
                "comprimento",
            ]
            linha = [
                self.var_status.get().strip() or "ativo",
                self.var_id.get().strip(),
                self.var_nome_produto.get().strip(),
                self.var_categoria.get().strip(),
                self.var_subcategoria.get().strip(),
                self.var_preco.get().strip(),
                self.var_destaque.get().strip(),
                descricao,
                campo_imagens,
                self.var_nome_variacao_1.get().strip(),
                self.var_variacoes_1.get().strip(),
                self.var_nome_variacao_2.get().strip(),
                self.var_variacoes_2.get().strip(),
                self.var_peso.get().strip(),
                self.var_altura.get().strip(),
                self.var_largura.get().strip(),
                self.var_comprimento.get().strip(),
            ]

            arquivo_existe = caminho_csv.exists()

            with open(caminho_csv, "a", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(
                    f,
                    delimiter=",",
                    quotechar='"',
                    quoting=csv.QUOTE_ALL,
                )
                if not arquivo_existe:
                    writer.writerow(cabecalho)
                writer.writerow(linha)

            self.gerar_linha_csv()
            self.carregar_produtos_csv()
            self.gerar_id_automatico_e_pasta()
            self._set_status("Produto adicionado ao produtos.csv com sucesso.")
            messagebox.showinfo("Sucesso", f"Produto adicionado em:\n{caminho_csv}")
        except Exception as e:
            messagebox.showerror("Erro", str(e))

    def carregar_produtos_csv(self):
        try:
            pasta_projeto = self.var_pasta_projeto.get().strip()
            if not pasta_projeto:
                return

            caminho_csv = self.obter_caminho_csv()
            if not caminho_csv.exists():
                self.produtos_csv_cache = []
                self.combo_produtos["values"] = []
                return

            produtos = []
            with open(caminho_csv, "r", newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    produtos.append(row)

            self.produtos_csv_cache = produtos
            opcoes = [
                f'{p.get("id", "").strip()} - {p.get("nome", "").strip()}'
                for p in produtos
            ]
            self.combo_produtos["values"] = opcoes
            self._set_status(f"{len(produtos)} produto(s) carregado(s) do CSV.")
        except Exception as e:
            messagebox.showerror("Erro", str(e))

    def carregar_produto_existente(self, _event=None):
        try:
            selecionado = self.var_produto_existente.get().strip()
            if not selecionado:
                return

            id_str = selecionado.split(" - ", 1)[0].strip()
            produto = None
            for item in self.produtos_csv_cache:
                if str(item.get("id", "")).strip() == id_str:
                    produto = item
                    break

            if not produto:
                return

            self.var_status.set(produto.get("status", "ativo").strip() or "ativo")
            self.var_id.set(produto.get("id", "").strip())
            self.var_nome_produto.set(produto.get("nome", "").strip())
            self.var_categoria.set(produto.get("categoria", "").strip() or "Chaveiro")
            self._atualizar_subcategorias()
            self.var_subcategoria.set(produto.get("subcategoria", "").strip())
            self.var_preco.set(produto.get("preco", "").strip())
            self.var_destaque.set(produto.get("destaque", "").strip())

            self.var_nome_variacao_1.set(produto.get("nome_variacao_1", "").strip())
            self.var_variacoes_1.set(produto.get("variacoes_1", "").strip())
            self.var_nome_variacao_2.set(produto.get("nome_variacao_2", "").strip())
            self.var_variacoes_2.set(produto.get("variacoes_2", "").strip())
            self.var_peso.set(produto.get("peso", "").strip())
            self.var_altura.set(produto.get("altura", "").strip())
            self.var_largura.set(produto.get("largura", "").strip())
            self.var_comprimento.set(produto.get("comprimento", "").strip())

            self.txt_descricao.delete("1.0", tk.END)
            self.txt_descricao.insert("1.0", produto.get("descricao", "").strip())

            self.var_pasta_produto.set(
                self.montar_nome_pasta_produto(
                    self.var_id.get().strip(),
                    self.var_nome_produto.get().strip()
                )
            )
            self._set_texto(self.txt_campo_imagens, produto.get("imagens", "").strip())
            self._set_texto(self.txt_linha_csv, "")
            self._set_texto(self.txt_arquivos_gerados, "")
            self._set_texto(self.txt_pasta_saida, "")

            self.imagens_selecionadas = []
            self.atualizar_lista_imagens()

            self._set_status("Produto carregado do CSV.")
        except Exception as e:
            messagebox.showerror("Erro", str(e))

    def ler_planilha_massa(self, caminho_arquivo):
        caminho = Path(str(caminho_arquivo).strip())
        extensao = caminho.suffix.strip().lower()

        if not str(caminho).strip():
            raise ValueError("Selecione uma planilha para importação em massa.")

        if not caminho.exists():
            raise ValueError(f"Arquivo não encontrado:\n{caminho}")

        if extensao == ".csv":
            with open(caminho, "r", newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                return [dict(row) for row in reader]

        if extensao == ".xlsx":
            try:
                wb = load_workbook(filename=str(caminho), data_only=True)
            except Exception as e:
                raise ValueError(
                    f"Não foi possível abrir o arquivo Excel.\n\n"
                    f"Arquivo: {caminho.name}\n"
                    f"Erro: {e}"
                ) from e

            nome_aba = "cadastro-em-massa"

            if nome_aba not in wb.sheetnames:
                raise ValueError(
                    f'A planilha Excel não contém a guia "{nome_aba}".\n\n'
                    f"Guias encontradas:\n" + "\n".join(wb.sheetnames)
                )

            ws = wb[nome_aba]

            linhas = list(ws.iter_rows(values_only=True))
            if not linhas:
                return []

            cabecalhos = [
                str(col).strip() if col is not None else ""
                for col in linhas[0]
            ]

            if not cabecalhos or cabecalhos[0].strip().lower() != "status":
                raise ValueError('A primeira coluna da planilha deve ser "status".')

            dados = []
            for linha in linhas[1:]:
                if linha is None:
                    continue

                item = {}
                linha_vazia = True

                for i, valor in enumerate(linha):
                    if i >= len(cabecalhos):
                        continue

                    chave = cabecalhos[i]
                    if not chave:
                        continue

                    valor_final = "" if valor is None else str(valor).strip()
                    if valor_final != "":
                        linha_vazia = False

                    item[chave] = valor_final

                if not linha_vazia:
                    dados.append(item)

            return dados

        raise ValueError(
            f"Formato não suportado.\n\n"
            f"Arquivo selecionado: {caminho.name}\n"
            f"Extensão encontrada: {extensao or '(sem extensão)'}\n\n"
            f"Use .csv ou .xlsx"
        )

    def processar_em_massa(self):
        try:
            pasta_projeto = self.var_pasta_projeto.get().strip()
            planilha = self.var_planilha_massa.get().strip()
            pasta_fotos = self.var_pasta_fotos.get().strip()

            if not pasta_projeto:
                raise ValueError("Selecione a pasta do projeto.")
            if not planilha:
                raise ValueError("Selecione a planilha CSV ou Excel.")
            if not pasta_fotos:
                raise ValueError("Selecione a pasta base das fotos.")

            caminho_csv_saida = self.obter_caminho_csv()

            produtos_saida = []
            relatorio = []

            linhas_planilha = self.ler_planilha_massa(planilha)

            if not linhas_planilha:
                raise ValueError("A planilha está vazia.")

            colunas_necessarias = {
                "status",
                "id",
                "nome",
                "categoria",
                "subcategoria",
                "preco",
                "destaque",
                "descricao",
                "nome_variacao_1",
                "variacoes_1",
                "nome_variacao_2",
                "variacoes_2",
                "peso",
                "altura",
                "largura",
                "comprimento",
                "caminho_origem",
                "pasta_produto",
            }
            colunas_planilha = set(linhas_planilha[0].keys())
            faltando = colunas_necessarias - colunas_planilha
            if faltando:
                raise ValueError(
                    "A planilha está sem as colunas obrigatórias:\n"
                    + ", ".join(sorted(faltando))
                )

            for numero_linha, row in enumerate(linhas_planilha, start=2):
                status = row.get("status", "").strip()
                if not valor_status_ativo(status):
                    relatorio.append(
                        f"Linha {numero_linha}: produto ignorado por status diferente de 'ativo'."
                    )
                    continue

                nome = row.get("nome", "").strip()
                categoria = row.get("categoria", "").strip()
                subcategoria = row.get("subcategoria", "").strip()
                destaque = row.get("destaque", "").strip()
                descricao = limpar_descricao_linha_unica(row.get("descricao", ""))
                preco = row.get("preco", "").strip()
                id_produto = row.get("id", "").strip()

                nome_variacao_1 = row.get("nome_variacao_1", "").strip()
                variacoes_1 = row.get("variacoes_1", "").strip()
                nome_variacao_2 = row.get("nome_variacao_2", "").strip()
                variacoes_2 = row.get("variacoes_2", "").strip()
                peso = row.get("peso", "").strip()
                altura = row.get("altura", "").strip()
                largura = row.get("largura", "").strip()
                comprimento = row.get("comprimento", "").strip()

                caminho_origem = row.get("caminho_origem", "").strip()
                pasta_produto = row.get("pasta_produto", "").strip()

                if not nome:
                    relatorio.append(f"Linha {numero_linha}: produto sem nome. Ignorado.")
                    continue

                if not pasta_produto:
                    pasta_produto = self.montar_nome_pasta_produto(id_produto, nome)

                if not caminho_origem:
                    relatorio.append(f"Linha {numero_linha}: {nome} sem caminho_origem. Ignorado.")
                    continue

                origem = Path(pasta_fotos) / Path(caminho_origem)

                if not origem.exists():
                    relatorio.append(f"Linha {numero_linha}: pasta não encontrada -> {origem}")
                    continue

                if not origem.is_dir():
                    relatorio.append(f"Linha {numero_linha}: caminho_origem não é pasta -> {origem}")
                    continue

                subpasta_categoria = self._obter_subpasta_categoria(categoria)

                destino = self.obter_base_imagens_backend() / subpasta_categoria / pasta_produto
                destino.mkdir(parents=True, exist_ok=True)

                arquivos_validos = sorted(
                    [
                        arq for arq in origem.iterdir()
                        if arq.is_file() and arq.suffix.lower() in {".png", ".jpg", ".jpeg", ".webp"}
                    ],
                    key=lambda x: x.name.lower()
                )

                if not arquivos_validos:
                    relatorio.append(f"Linha {numero_linha}: sem imagens válidas em {origem}")
                    continue

                imagens = []
                arquivos_copiados = []

                for contador, arquivo in enumerate(arquivos_validos, start=1):
                    ext = arquivo.suffix.lower()
                    novo_nome = f"{contador}{ext}"
                    destino_arquivo = destino / novo_nome

                    shutil.copy2(arquivo, destino_arquivo)

                    caminho_backend = self.montar_caminho_relativo_imagem_backend(
                        subpasta_categoria,
                        pasta_produto,
                        novo_nome,
                    )

                    imagens.append(caminho_backend)
                    arquivos_copiados.append(f"{arquivo.name} -> {novo_nome}")

                produtos_saida.append([
                    "ativo",
                    id_produto,
                    nome,
                    categoria,
                    subcategoria,
                    preco,
                    destaque,
                    descricao,
                    "|".join(imagens),
                    nome_variacao_1,
                    variacoes_1,
                    nome_variacao_2,
                    variacoes_2,
                    peso,
                    altura,
                    largura,
                    comprimento,
                ])

                relatorio.append(
                    f"[OK] {nome}\n"
                    f"Origem: {origem}\n"
                    f"Destino: {destino}\n"
                    f"Arquivos: {len(arquivos_copiados)}\n"
                )

            with open(caminho_csv_saida, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(
                    f,
                    delimiter=",",
                    quotechar='"',
                    quoting=csv.QUOTE_ALL,
                )

                writer.writerow([
                    "status",
                    "id",
                    "nome",
                    "categoria",
                    "subcategoria",
                    "preco",
                    "destaque",
                    "descricao",
                    "imagens",
                    "nome_variacao_1",
                    "variacoes_1",
                    "nome_variacao_2",
                    "variacoes_2",
                    "peso",
                    "altura",
                    "largura",
                    "comprimento",
                ])
                writer.writerows(produtos_saida)

            self.carregar_produtos_csv()
            self._set_texto(self.txt_arquivos_gerados, "\n\n".join(relatorio))
            self._set_texto(self.txt_pasta_saida, str(caminho_csv_saida))
            self._set_texto(self.txt_campo_imagens, "")
            self._set_texto(self.txt_linha_csv, "")

            self._set_status(f"{len(produtos_saida)} produto(s) processado(s) com sucesso.")
            messagebox.showinfo(
                "Sucesso",
                f"Importação em massa concluída.\n\n"
                f"{len(produtos_saida)} produto(s) processado(s).\n\n"
                f"CSV gerado em:\n{caminho_csv_saida}"
            )

        except Exception as e:
            messagebox.showerror("Erro", str(e))

    def copiar_campo_imagens(self):
        texto = self.txt_campo_imagens.get("1.0", tk.END).strip()
        if not texto:
            messagebox.showwarning("Aviso", 'O campo "imagens" está vazio.')
            return
        self.root.clipboard_clear()
        self.root.clipboard_append(texto)
        self._set_status('Campo "imagens" copiado.')

    def copiar_linha_csv(self):
        texto = self.txt_linha_csv.get("1.0", tk.END).strip()
        if not texto:
            messagebox.showwarning("Aviso", "A linha CSV está vazia.")
            return
        self.root.clipboard_clear()
        self.root.clipboard_append(texto)
        self._set_status("Linha CSV copiada.")

    def abrir_pasta_produto(self):
        try:
            destino = self._obter_pasta_destino()
            abrir_no_sistema(destino)
            self._set_status("Pasta do produto aberta.")
        except Exception as e:
            messagebox.showerror("Erro", str(e))

    def abrir_csv(self):
        try:
            if not self.var_pasta_projeto.get().strip():
                raise ValueError("Selecione a pasta do projeto primeiro.")
            caminho_csv = self.obter_caminho_csv()
            abrir_no_sistema(caminho_csv)
            self._set_status("produtos.csv aberto.")
        except Exception as e:
            messagebox.showerror("Erro", str(e))

    def limpar_tudo(self):
        self.var_produto_existente.set("")
        self.var_status.set("ativo")
        self.var_id.set("")
        self.var_nome_produto.set("")
        self.var_categoria.set("Chaveiro")
        self._atualizar_subcategorias()
        self.var_preco.set("")
        self.var_destaque.set("")
        self.var_pasta_produto.set("")

        self.var_nome_variacao_1.set("")
        self.var_variacoes_1.set("")
        self.var_nome_variacao_2.set("")
        self.var_variacoes_2.set("")
        self.var_peso.set("")
        self.var_altura.set("")
        self.var_largura.set("")
        self.var_comprimento.set("")

        self.txt_descricao.delete("1.0", tk.END)
        self.lista_imagens.delete(0, tk.END)
        self.imagens_selecionadas = []

        self._set_texto(self.txt_pasta_saida, "")
        self._set_texto(self.txt_campo_imagens, "")
        self._set_texto(self.txt_linha_csv, "")
        self._set_texto(self.txt_arquivos_gerados, "")

        self.gerar_id_automatico_e_pasta()
        self._set_status("Campos limpos.")


def main():
    root = tk.Tk()
    app = GeradorCatalogoApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()